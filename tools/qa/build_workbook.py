#!/usr/bin/env python3
"""
Build a standardized, tester-friendly QA workbook (.xlsx) for one module: 5 sheets in ONE file —
`00 Tổng quan` · `PCL` · `Test_Cases` · `Test_Run` · `Bugs`.

Format chuẩn cho tester:
- `00 Tổng quan`: tiêu đề + bảng coverage + **Legend** (màu Priority, màu Kết quả, ý nghĩa tag, hướng dẫn).
- `Test_Cases`: master (sinh từ JSON, không sửa tay); ô **Ưu tiên** tô màu; dropdown `Loại`/`Ưu tiên` (phòng hờ).
- `Test_Run`: **append-only**; ô **Kết quả** có **dropdown** + **tự tô màu**; tách cột `Vòng` và `Build/Version`.
- `Bugs`: dropdown cho Mức độ / Ưu tiên / Trạng thái. (Khuyến nghị log bug ở issue tracker.)

Input  : JSON (xem tools/README.md): { title, filename, app, module, source,
          test_cases:[{ id, area, feature, precond, data, steps[], expected[], type, priority, notes }], pcl?:[...] }
          - `pcl` (optional) = checklist thủ công; mỗi PCL có `covered_by` liệt kê TC ID → map ngược về Test_Cases.
Usage  : pip install openpyxl ;  python build_workbook.py <input.json> [output.xlsx]
"""
import json
import sys
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1E3D1A"
BAND = "F5F7F4"
PRIORITY_FILL = {"Critical": "F4CCCC", "High": "FCE5CD", "Medium": "FFF2CC", "Low": "EFEFEF"}
PRIORITY_FONT = {"Critical": "990000", "High": "B45F06", "Medium": "7F6000", "Low": "666666"}
RESULT_FILL = {"Pass": "D9EAD3", "Fail": "F4CCCC", "Blocked": "FCE5CD", "Skip": "EFEFEF", "Not Run": "FFFFFF"}
RESULT_FONT = {"Pass": "38761D", "Fail": "990000", "Blocked": "B45F06", "Skip": "666666", "Not Run": "666666"}
PRIORITY_RANK = {"Critical": 3, "High": 2, "Medium": 1, "Low": 0}

RESULTS = ["Pass", "Fail", "Blocked", "Skip", "Not Run"]
SEVERITIES = ["Critical", "Major", "Minor", "Trivial"]
PRIORITIES = ["Critical", "High", "Medium", "Low"]
TYPES = ["Functional", "Negative", "Boundary", "Integration", "Regression", "Online", "Offline", "Permission", "Sync/LWW"]
BUG_STATUS = ["Open", "In Progress", "Fixed", "Verified", "Closed", "Won't fix"]

TAG_LEGEND = [
    ("[AUTH]", "Đăng nhập / token / phiên"), ("[VALIDATION]", "Kiểm tra ràng buộc nhập"),
    ("[ERROR]", "Xử lý lỗi / thông báo"), ("[OFFLINE]", "Hành vi offline"),
    ("[PERMISSION]", "Phân quyền / scope"), ("[SYNC]", "Đồng bộ / LWW"),
    ("[NAV]", "Điều hướng"), ("[UI]", "Hiển thị / giao diện"), ("[EDGE]", "Edge case / concurrency"),
]

TEST_CASE_COLS = ["No.", "TC ID", "PCL ID", "Nhóm", "Chức năng", "Tiền điều kiện",
                  "Dữ liệu test", "Các bước", "Kết quả mong đợi", "Loại", "Ưu tiên", "Nguồn", "Ghi chú"]
TEST_CASE_W = [5, 18, 16, 20, 26, 26, 20, 46, 46, 13, 11, 16, 26]
TEST_CASE_WRAP = [False, False, False, True, True, True, True, True, True, False, False, False, True]
PCL_COLS = ["PCL ID", "Khu vực", "Check Point", "Nguồn", "Loại", "Ưu tiên", "Covered By", "Ghi chú"]
PCL_W = [18, 20, 44, 16, 13, 11, 30, 22]
PCL_WRAP = [False, True, True, False, False, False, True, True]
RUN_COLS = ["Run ID", "TC ID", "Vòng", "Build/Version", "Môi trường", "Tester", "Ngày",
            "Kết quả", "Actual Result", "Bug ID", "Ghi chú"]
RUN_W = [16, 18, 7, 14, 12, 14, 12, 12, 34, 14, 24]
RUN_WRAP = [False, False, False, False, False, False, False, False, True, False, True]
BUG_COLS = ["Bug ID", "TC ID", "Tiêu đề", "Mức độ", "Ưu tiên", "Trạng thái", "Bước tái hiện",
            "Mong đợi", "Thực tế", "Môi trường", "Người báo", "Phụ trách", "Ghi chú"]
BUG_W = [12, 18, 30, 11, 11, 14, 36, 28, 28, 12, 14, 14, 22]
BUG_WRAP = [False, False, True, False, False, False, True, True, True, False, False, False, True]

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor=PRIMARY)
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _t(v):
    return "\n".join(str(x) for x in v) if isinstance(v, list) else ("" if v is None else str(v))


def _derive_pcl(spec, cases):
    """Auto: 1 PCL checkpoint per `area`, covered_by = TC IDs of that area."""
    app = spec.get("app", "APP")
    module = spec.get("module", "MODULE").upper().replace(" ", "-")
    areas = []
    for c in cases:
        if c.get("area", "General") not in areas:
            areas.append(c.get("area", "General"))
    pcl, tc_to_pcl = [], {}
    for i, area in enumerate(areas, start=1):
        pid = f"{app}-{module}-PCL-{i:02d}"
        group = [c for c in cases if c.get("area", "General") == area]
        for c in group:
            tc_to_pcl[c["id"]] = pid
        types = sorted({c.get("type", "") for c in group if c.get("type")})
        prio = max((c.get("priority", "Low") for c in group), key=lambda p: PRIORITY_RANK.get(p, 0))
        pcl.append({"pcl_id": pid, "area": area, "check_point": f"Kiểm thử nhóm: {area}",
                    "source": spec.get("source", ""), "type": "Mixed" if len(types) > 1 else (types[0] if types else ""),
                    "priority": prio, "covered_by": ", ".join(c["id"] for c in group), "notes": ""})
    return pcl, tc_to_pcl


def _header(ws, cols, widths, freeze=None):
    for ci, (h, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30
    if freeze:
        ws.freeze_panes = freeze


def _rows(ws, rows, wrap, band=True):
    for ri, row in enumerate(rows, start=2):
        lines = 1
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = LEFT if wrap[ci - 1] else CENTER
            c.border = BORDER
            if band and ri % 2 == 1:
                c.fill = PatternFill("solid", fgColor=BAND)
            if isinstance(val, str):
                lines = max(lines, val.count("\n") + 1)
        ws.row_dimensions[ri].height = max(30, lines * 15)


def _dropdown(ws, col, options):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}1000")


def build(spec, out_path):
    cases = spec.get("test_cases", [])
    source = spec.get("source", "")
    if spec.get("pcl"):
        pcl_rows = spec["pcl"]
        tc_to_pcl = {}
        for p in pcl_rows:
            for tc in [x.strip() for x in str(p.get("covered_by", "")).split(",") if x.strip()]:
                tc_to_pcl[tc] = p["pcl_id"]
    else:
        pcl_rows, tc_to_pcl = _derive_pcl(spec, cases)

    by_type = {}
    for c in cases:
        by_type[c.get("type", "?")] = by_type.get(c.get("type", "?"), 0) + 1
    n_assume = sum(1 for c in cases if "[ASSUMPTION]" in (c.get("feature", "") + c.get("notes", "")))
    n_pending = sum(1 for c in cases if "[Pending" in (c.get("feature", "") + c.get("notes", "")))

    wb = Workbook()
    wb.remove(wb.active)

    # ── 00 Tổng quan ──
    ov = wb.create_sheet("00 Tổng quan")
    ov.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ov.column_dimensions[col].width = 16
    ov.merge_cells("A1:F1")
    t = ov["A1"]; t.value = spec.get("title", "Test Cases")
    t.font = Font(bold=True, size=15, color="FFFFFF"); t.fill = HDR_FILL; t.alignment = Alignment(vertical="center")
    ov.row_dimensions[1].height = 32
    ov["A2"] = f"App: {spec.get('app','')}    Module: {spec.get('module','')}    Source: {source}"
    ov["A2"].font = Font(color="555555")
    ov["A3"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ov["A3"].font = Font(italic=True, color="888888")

    r = 5
    ov.cell(r, 1, "Coverage").font = Font(bold=True, size=12); r += 1
    for label, val in [("Tổng số test case", len(cases))] + sorted(by_type.items()) + [("[ASSUMPTION]", n_assume), ("[Pending]", n_pending)]:
        a = ov.cell(r, 1, label); b = ov.cell(r, 2, val)
        a.border = BORDER; b.border = BORDER; b.alignment = CENTER; a.fill = PatternFill("solid", fgColor=BAND)
        r += 1
    r += 1
    ov.cell(r, 1, "Chú giải (Legend)").font = Font(bold=True, size=12); r += 1
    ov.cell(r, 1, "Mức ưu tiên:").font = Font(bold=True)
    for i, p in enumerate(PRIORITIES):
        c = ov.cell(r, 2 + i, p); c.fill = PatternFill("solid", fgColor=PRIORITY_FILL[p])
        c.font = Font(color=PRIORITY_FONT[p], bold=True); c.alignment = CENTER; c.border = BORDER
    r += 1
    ov.cell(r, 1, "Kết quả test:").font = Font(bold=True)
    for i, res in enumerate(RESULTS):
        c = ov.cell(r, 2 + i, res); c.fill = PatternFill("solid", fgColor=RESULT_FILL[res])
        c.font = Font(color=RESULT_FONT[res], bold=True); c.alignment = CENTER; c.border = BORDER
    r += 2
    ov.cell(r, 1, "Tag (trong Ghi chú):").font = Font(bold=True); r += 1
    for tag, mean in TAG_LEGEND:
        ov.cell(r, 1, tag).font = Font(bold=True, color="1E3D1A")
        ov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ov.cell(r, 2, mean); r += 1
    r += 1
    ov.cell(r, 1, "Hướng dẫn:").font = Font(bold=True); r += 1
    for line in [
        "• Test_Cases = bản gốc (master) — KHÔNG sửa tay; muốn đổi thì sửa file JSON rồi generate lại.",
        "• Test_Run = mỗi lần test 1 case thêm 1 DÒNG MỚI (chọn Kết quả từ dropdown). Nhiều người/nhiều vòng đều ghi ở đây.",
        "• Bugs = chọn Mức độ/Ưu tiên/Trạng thái từ dropdown; nếu dùng issue tracker thì ghi link vào Bug ID.",
    ]:
        ov.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ov.cell(r, 1, line).alignment = Alignment(wrap_text=True, vertical="top"); r += 1

    # ── PCL ──
    ws = wb.create_sheet("PCL")
    _header(ws, PCL_COLS, PCL_W, freeze="A2")
    _rows(ws, [[p["pcl_id"], p["area"], p["check_point"], p.get("source", source), p.get("type", ""),
                p.get("priority", ""), p.get("covered_by", ""), p.get("notes", "")] for p in pcl_rows], PCL_WRAP)

    # ── Test_Cases ──
    ws = wb.create_sheet("Test_Cases")
    _header(ws, TEST_CASE_COLS, TEST_CASE_W, freeze="C2")
    _rows(ws, [[i, c.get("id", ""), tc_to_pcl.get(c.get("id", ""), ""), c.get("area", ""), c.get("feature", ""),
                _t(c.get("precond", "")), _t(c.get("data", "")), _t(c.get("steps", "")), _t(c.get("expected", "")),
                c.get("type", ""), c.get("priority", ""), c.get("source", source), c.get("notes", "")]
               for i, c in enumerate(cases, start=1)], TEST_CASE_WRAP, band=False)
    for ri, c in enumerate(cases, start=2):
        p = c.get("priority", "")
        if p in PRIORITY_FILL:
            cell = ws.cell(ri, 11)
            cell.fill = PatternFill("solid", fgColor=PRIORITY_FILL[p]); cell.font = Font(color=PRIORITY_FONT[p], bold=True)
    _dropdown(ws, "J", TYPES)       # Loại
    _dropdown(ws, "K", PRIORITIES)  # Ưu tiên

    # ── Test_Run (dropdown + conditional color, Kết quả = cột H) ──
    ws = wb.create_sheet("Test_Run")
    _header(ws, RUN_COLS, RUN_W, freeze="A2")
    first = cases[0]["id"] if cases else ""
    _rows(ws, [["RUN-YYYYMMDD-01", first, "1", "v1.0.0", "Staging", "Tên QA", "YYYY-MM-DD",
                "Not Run", "", "", "Mỗi lần chạy thêm 1 dòng mới"]], RUN_WRAP)
    _dropdown(ws, "H", RESULTS)
    for res in RESULTS:
        ws.conditional_formatting.add("H2:H1000", CellIsRule(
            operator="equal", formula=[f'"{res}"'],
            fill=PatternFill("solid", fgColor=RESULT_FILL[res]), font=Font(color=RESULT_FONT[res], bold=True)))

    # ── Bugs (dropdowns) ──
    ws = wb.create_sheet("Bugs")
    _header(ws, BUG_COLS, BUG_W, freeze="A2")
    _rows(ws, [["BUG-01", first, "Tiêu đề bug", "Major", "High", "Open",
                "1. … | 2. …", "Mong đợi", "Thực tế", "Staging", "Tên QA", "Tên Dev", ""]], BUG_WRAP)
    _dropdown(ws, "D", SEVERITIES); _dropdown(ws, "E", PRIORITIES); _dropdown(ws, "F", BUG_STATUS)

    wb.save(out_path)
    print(f"Saved {out_path}  (5 sheets, {len(cases)} test cases)")


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python build_workbook.py <input.json> [output.xlsx]")
    spec = json.load(open(sys.argv[1], encoding="utf-8"))
    out = sys.argv[2] if len(sys.argv) > 2 else spec.get("filename") or f"{spec.get('title', 'TestCases')}.xlsx"
    build(spec, out)


if __name__ == "__main__":
    main()
